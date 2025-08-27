import json
import os
import argparse
from collections import defaultdict
from openpyxl import Workbook
from google.cloud import storage
import io
import re

def parse_tpu_type(tpu_type):
    """
    Parses the TPU type string to extract chip count.

    Args:
        tpu_type (str): The TPU type string, e.g., "v5p-128", "v5p-256".

    Returns:
        dict: Containing "chips".

    Raises:
        ValueError: If the chip count cannot be extracted.
    """
    match = re.search(r'(\d+)$', tpu_type)
    if not match:
        raise ValueError(f"Cannot extract chip count from tpu_type: {tpu_type}")

    chips = int(match.group(1))
    return {
        "chips": chips
    }

def download_gcs_blob_as_text(gcs_path):
    try:
        if not gcs_path.startswith("gs://"):
            raise ValueError("GCS path must start with gs://")
        parts = gcs_path[5:].split("/", 1)
        bucket_name = parts[0]
        blob_name = parts[1]
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        return blob.download_as_text()
    except Exception as e:
        raise

def upload_to_gcs(bucket_name, blob_name, data):
    try:
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        blob.upload_from_file(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        raise

def generate_excel_report(jsonl_gcs_path, tpu_type, excel_gcs_path):
    try:
        config = parse_tpu_type(tpu_type)
    except ValueError as e:
        return

    chip_count = config["chips"]

    data_by_test = defaultdict(lambda: defaultdict(dict))

    def process_jsonl(jsonl_content):
        if not jsonl_content: return

        for line in jsonl_content.strip().split('\n'):
            if not line: continue
            try:
                record = json.loads(line)
                if 'metrics' in record and 'dimensions' in record:
                    metrics = record['metrics']
                    dims = record['dimensions']
                    test_name = dims.get('test_name')
                    matrix_dim = dims.get('matrix_dim')

                    if test_name and matrix_dim is not None:
                        try:
                            dim = int(matrix_dim)
                            data_by_test[test_name][dim][chip_count] = metrics
                        except ValueError:
                            pass
            except json.JSONDecodeError:
                pass
            except Exception as e:
                pass
    try:
        jsonl_content = download_gcs_blob_as_text(jsonl_gcs_path)
        process_jsonl(jsonl_content)
    except Exception as e:
        return

    if not data_by_test:
        return

    wb = Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])

    metrics_keys = ["ici_bandwidth_gbyte_s_p50", "ici_bandwidth_gbyte_s_p90", "ici_bandwidth_gbyte_s_p95", "ici_bandwidth_gbyte_s_p99", "ici_bandwidth_gbyte_s_avg"]

    for test_name in sorted(data_by_test.keys()):
        matrix_data = data_by_test[test_name]
        safe_test_name = "".join(c for c in test_name if c.isalnum() or c in (' ', '_')).rstrip()[:31]
        ws = wb.create_sheet(title=safe_test_name)
        matrix_dims = sorted(matrix_data.keys())

        current_col = 1
        for metric in metrics_keys:
            ws.cell(row=1, column=current_col, value=metric)
            ws.cell(row=2, column=current_col, value="dimensions\\TPUs")
            ws.cell(row=2, column=current_col + 1, value=chip_count)

            for row_idx, dim in enumerate(matrix_dims):
                ws.cell(row=3 + row_idx, column=current_col, value=dim)
                core_metrics = matrix_data[dim].get(chip_count)
                metric_val = core_metrics.get(metric) if core_metrics else ""
                ws.cell(row=3 + row_idx, column=current_col + 1, value=metric_val)

            current_col += 3

    try:
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        if not excel_gcs_path.startswith("gs://"):
            raise ValueError("GCS output path must start with gs://")
        parts = excel_gcs_path[5:].split("/", 1)
        bucket_name = parts[0]
        blob_name = parts[1]
        upload_to_gcs(bucket_name, blob_name, excel_buffer)
    except Exception as e:
        raise

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Generate benchmark report for a single TPU type and upload to GCS.")
    parser.add_argument("--gcs_path", required=True, help="GCS path to metrics_report.jsonl file")
    parser.add_argument("--tpu_type", required=True, help="TPU type (e.g., v5p-128, v5p-256)")
    parser.add_argument("--gcs_output_path", required=True, help="GCS path to save the generated Excel file")
    args = parser.parse_args()
    try:
        generate_excel_report(args.gcs_path, args.tpu_type, args.gcs_output_path)
    except Exception as e:
        raise
